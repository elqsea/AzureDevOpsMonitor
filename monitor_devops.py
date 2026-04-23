import requests
import json
import time
import base64
import os
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── Config ───────────────────────────────────────────────────────────────────
CONFIG_FILE = os.path.join(os.path.dirname(__file__), "config.json")

with open(CONFIG_FILE, "r", encoding="utf-8") as f:
    config = json.load(f)

PAT               = config["pat_token"]
ORGANIZATION      = config["organization"]
PROJECT           = config["project"]
INTERVAL_MINUTES  = config.get("interval_minutes", 30)
OUTPUT_DIR        = os.path.join(os.path.dirname(__file__), config.get("output_dir", "output"))
RCA_KEYWORDS      = [k.lower() for k in config.get("rca_field_keywords", ["rca", "root cause", "causa raiz"])]
TIMELINE_KEYWORDS = [k.lower() for k in config.get("timeline_field_keywords", ["timeline", "cronologia"])]
ITEM_TYPES        = config.get("work_item_types", [])

API_BASE = f"https://dev.azure.com/{ORGANIZATION}/{PROJECT}/_apis"

# ─── Auth ─────────────────────────────────────────────────────────────────────
def _headers():
    token = base64.b64encode(f":{PAT}".encode()).decode()
    return {"Authorization": f"Basic {token}", "Content-Type": "application/json"}

# ─── API calls ────────────────────────────────────────────────────────────────
def fetch_all_ids():
    type_filter = ""
    if ITEM_TYPES:
        types_str = ", ".join(f"'{t}'" for t in ITEM_TYPES)
        type_filter = f" AND [System.WorkItemType] IN ({types_str})"

    wiql = {
        "query": (
            f"SELECT [System.Id] FROM WorkItems "
            f"WHERE [System.TeamProject] = '{PROJECT}'{type_filter} "
            f"ORDER BY [System.ChangedDate] DESC"
        )
    }
    url = f"{API_BASE}/wit/wiql?api-version=7.2-preview.2"
    r = requests.post(url, headers=_headers(), json=wiql, timeout=30)
    r.raise_for_status()
    return [item["id"] for item in r.json().get("workItems", [])]


def fetch_items_batch(ids):
    """Fetch work item details in batches of 200."""
    results = []
    batch_size = 200
    for i in range(0, len(ids), batch_size):
        batch = ids[i : i + batch_size]
        ids_str = ",".join(map(str, batch))
        url = (
            f"{API_BASE}/wit/workitems"
            f"?ids={ids_str}&$expand=All&api-version=7.2-preview.3"
        )
        r = requests.get(url, headers=_headers(), timeout=30)
        r.raise_for_status()
        results.extend(r.json().get("value", []))
    return results


def fetch_item_comments(item_id):
    """Fetch comments for a work item."""
    url = f"{API_BASE}/wit/workitems/{item_id}/comments?api-version=7.2-preview.4"
    try:
        r = requests.get(url, headers=_headers(), timeout=10)
        if r.status_code == 200:
            return [c.get("text", "") for c in r.json().get("comments", [])]
    except Exception:
        pass
    return []

# ─── Extraction helpers ───────────────────────────────────────────────────────
def _strip_html(html_text):
    """Remove HTML tags for plain-text comparison."""
    if not html_text:
        return ""
    return re.sub(r"<[^>]+>", " ", str(html_text)).strip()


def _search_text_block(text, keywords):
    """Return the first paragraph that contains any keyword."""
    text_lower = text.lower()
    for kw in keywords:
        idx = text_lower.find(kw)
        if idx != -1:
            # grab surrounding context (up to 500 chars)
            start = max(0, idx - 50)
            end = min(len(text), idx + 500)
            return text[start:end].strip()
    return ""


def extract_rca_timeline(fields, comments):
    rca = ""
    timeline = ""

    # 1. Check every field name
    for field_name, value in fields.items():
        if value is None:
            continue
        fn_lower = field_name.lower()
        val_str = _strip_html(value) if isinstance(value, str) else str(value)

        if not rca and any(k in fn_lower for k in RCA_KEYWORDS):
            rca = val_str
        if not timeline and any(k in fn_lower for k in TIMELINE_KEYWORDS):
            timeline = val_str

    # 2. Fallback: search in description text
    description = _strip_html(fields.get("System.Description", ""))
    if not rca:
        rca = _search_text_block(description, RCA_KEYWORDS)
    if not timeline:
        timeline = _search_text_block(description, TIMELINE_KEYWORDS)

    # 3. Fallback: search in comments
    for comment in comments:
        comment_plain = _strip_html(comment)
        if not rca:
            rca = _search_text_block(comment_plain, RCA_KEYWORDS)
        if not timeline:
            timeline = _search_text_block(comment_plain, TIMELINE_KEYWORDS)
        if rca and timeline:
            break

    return rca, timeline

# ─── Processing ───────────────────────────────────────────────────────────────
def process_items(raw_items):
    processed = []
    for item in raw_items:
        fields = item.get("fields", {})
        relations = item.get("relations", [])

        child_ids = [
            int(rel["url"].split("/")[-1])
            for rel in relations
            if rel.get("rel") == "System.LinkTypes.Hierarchy-Forward"
        ]

        parent_id = next(
            (int(rel["url"].split("/")[-1])
             for rel in relations
             if rel.get("rel") == "System.LinkTypes.Hierarchy-Reverse"),
            None
        )

        comments = fetch_item_comments(item["id"])
        rca, timeline = extract_rca_timeline(fields, comments)

        assigned = fields.get("System.AssignedTo", "")
        if isinstance(assigned, dict):
            assigned = assigned.get("displayName", "")

        # Collect all scalar custom fields for HTML detail view
        custom_fields = {
            k: v
            for k, v in fields.items()
            if isinstance(v, (str, int, float, bool, type(None)))
            and k not in (
                "System.Description",
                "System.TeamProject",
                "Microsoft.VSTS.Common.StateChangeDate",
            )
        }

        processed.append({
            "id":           item["id"],
            "title":        fields.get("System.Title", ""),
            "type":         fields.get("System.WorkItemType", ""),
            "state":        fields.get("System.State", ""),
            "assigned_to":  assigned,
            "created_date": fields.get("System.CreatedDate", ""),
            "changed_date": fields.get("System.ChangedDate", ""),
            "description":  _strip_html(fields.get("System.Description", "")),
            "tags":         fields.get("System.Tags", ""),
            "rca":          rca,
            "timeline":     timeline,
            "parent_id":    parent_id,
            "child_ids":    child_ids,
            "comments":     [_strip_html(c) for c in comments],
            "url":          f"https://dev.azure.com/{ORGANIZATION}/{PROJECT}/_workitems/edit/{item['id']}",
            "custom_fields": custom_fields,
        })
    return processed

# ─── Save JSON ────────────────────────────────────────────────────────────────
def save_json(data):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    payload = {
        "last_updated": datetime.now().isoformat(),
        "organization":  ORGANIZATION,
        "project":       PROJECT,
        "total_items":   len(data),
        "items":         data,
    }
    # timestamped backup
    with open(os.path.join(OUTPUT_DIR, f"devops_{ts}.json"), "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, default=str)
    # always-fresh latest file
    latest = os.path.join(OUTPUT_DIR, "devops_latest.json")
    with open(latest, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, default=str)
    print(f"  JSON → {latest}")
    return payload

# ─── Save Excel ───────────────────────────────────────────────────────────────
def save_excel(data):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Work Items ──
    ws = wb.active
    ws.title = "Work Items"

    HEADERS = [
        "ID", "Tipo", "Título", "Estado", "Asignado a",
        "Fecha Creación", "Última Modificación", "RCA", "Timeline", "URL"
    ]
    HDR_FILL = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    HDR_FONT = Font(color="FFFFFF", bold=True)
    RCA_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    NO_RCA_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, hdr in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=hdr)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for row_idx, item in enumerate(data, 2):
        has_rca = bool(item["rca"] or item["timeline"])
        row_fill = RCA_FILL if has_rca else NO_RCA_FILL
        values = [
            item["id"], item["type"], item["title"], item["state"],
            item["assigned_to"], item["created_date"], item["changed_date"],
            item["rca"], item["timeline"], item["url"]
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.fill = row_fill
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = border

    # Column widths
    col_widths = [8, 14, 50, 15, 25, 22, 22, 60, 60, 50]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # ── Sheet 2: RCA Summary ──
    ws2 = wb.create_sheet("RCA & Timeline")
    ws2_hdrs = ["ID", "Tipo", "Título", "Estado", "RCA", "Timeline"]
    for col, hdr in enumerate(ws2_hdrs, 1):
        c = ws2.cell(row=1, column=col, value=hdr)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center")

    rca_items = [i for i in data if i["rca"] or i["timeline"]]
    for row_idx, item in enumerate(rca_items, 2):
        for col, val in enumerate(
            [item["id"], item["type"], item["title"], item["state"], item["rca"], item["timeline"]], 1
        ):
            c = ws2.cell(row=row_idx, column=col, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")

    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 50
    ws2.column_dimensions["D"].width = 15
    ws2.column_dimensions["E"].width = 70
    ws2.column_dimensions["F"].width = 70
    ws2.freeze_panes = "A2"

    path_ts = os.path.join(OUTPUT_DIR, f"devops_{ts}.xlsx")
    path_latest = os.path.join(OUTPUT_DIR, "devops_latest.xlsx")
    wb.save(path_ts)
    wb.save(path_latest)
    print(f"  Excel → {path_latest}")

# ─── Dashboard HTML ───────────────────────────────────────────────────────────
def save_dashboard(payload):
    template_path = os.path.join(os.path.dirname(__file__), "dashboard.html")
    if not os.path.exists(template_path):
        print("  dashboard.html not found – skipping HTML update")
        return

    with open(template_path, "r", encoding="utf-8") as f:
        template = f.read()

    json_blob = json.dumps(payload, ensure_ascii=False, default=str)
    # inject data between the sentinel markers
    updated = re.sub(
        r"(// DATA_START\n)[\s\S]*?(// DATA_END)",
        f"// DATA_START\nconst DEVOPS_DATA = {json_blob};\n// DATA_END",
        template,
    )
    out_path = os.path.join(OUTPUT_DIR, "dashboard.html")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(updated)
    print(f"  HTML  → {out_path}")

# ─── Main loop ────────────────────────────────────────────────────────────────
def run_once():
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"\n[{now}] Fetching work items from Azure DevOps…")
    ids = fetch_all_ids()
    print(f"  Found {len(ids)} work item(s)")
    if not ids:
        return

    raw = fetch_items_batch(ids)
    data = process_items(raw)

    rca_count = sum(1 for i in data if i["rca"] or i["timeline"])
    print(f"  Processed {len(data)} items | {rca_count} with RCA/Timeline")

    payload = save_json(data)
    save_excel(data)
    save_dashboard(payload)


def main():
    print("=" * 55)
    print(f"  Azure DevOps Monitor")
    print(f"  Org:      {ORGANIZATION}")
    print(f"  Project:  {PROJECT}")
    print(f"  Interval: {INTERVAL_MINUTES} min")
    print(f"  Output:   {OUTPUT_DIR}")
    print("=" * 55)

    while True:
        try:
            run_once()
        except Exception as exc:
            print(f"  ERROR: {exc}")

        next_run = datetime.now().strftime("%H:%M:%S")
        print(f"  Next run in {INTERVAL_MINUTES} min  (started {next_run})")
        time.sleep(INTERVAL_MINUTES * 60)


if __name__ == "__main__":
    main()
